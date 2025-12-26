"""
Export Router - Generate XLS, PDF and CSV exports for data tables
Enhanced with company header, logo, currency formatting, and footer totals
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
import io
import csv
import os
from datetime import datetime
from pathlib import Path

import database
import models
from models.models_inventory import Location, Warehouse
from models.models_receiving import InventoryBatch
from models import models_ledger
from auth import get_current_user

router = APIRouter(prefix="/export", tags=["Export"])

# Company info
COMPANY_NAME = "MATA RANTAI"
COMPANY_TAGLINE = "Supply Chain Management System"
LOGO_PATH = Path(__file__).parent.parent / "static" / "logo.png"


def format_currency(value, symbol="Rp", position="before", thousand_sep=".", decimal_sep=",", decimal_places=0):
    """Format number as currency using provided settings"""
    if value is None:
        return f"{symbol} 0" if position == "before" else f"0 {symbol}"
    try:
        # Format with decimal places
        dec_places = int(decimal_places) if isinstance(decimal_places, str) else decimal_places
        formatted = f"{value:,.{dec_places}f}"
        
        # Replace separators (Python uses comma for thousand, dot for decimal)
        if dec_places > 0:
            parts = formatted.split(".")
            parts[0] = parts[0].replace(",", thousand_sep)
            formatted = decimal_sep.join(parts)
        else:
            formatted = formatted.replace(",", thousand_sep).split(".")[0]
        
        # Apply position
        if position == "before":
            return f"{symbol} {formatted}"
        else:
            return f"{formatted} {symbol}"
    except:
        return f"{symbol} {value}"


async def get_tenant_currency_settings(db, tenant_id):
    """Get currency settings for a tenant"""
    from models.models_settings import TenantSettings
    result = await db.execute(
        select(TenantSettings).where(TenantSettings.tenant_id == tenant_id)
    )
    settings = result.scalar_one_or_none()
    
    if settings:
        return {
            "symbol": settings.currency_symbol or "Rp",
            "position": settings.currency_position or "before",
            "thousand_sep": settings.thousand_separator or ".",
            "decimal_sep": settings.decimal_separator or ",",
            "decimal_places": int(settings.decimal_places) if settings.decimal_places else 0
        }
    
    # Default settings
    return {
        "symbol": "Rp",
        "position": "before",
        "thousand_sep": ".",
        "decimal_sep": ",",
        "decimal_places": 0
    }


def create_csv_response(data: list, columns: list, filename: str, title: str = "", footer_data: dict = None):
    """Create CSV file from data with header and footer"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header - Logo placeholder on left, info on right (simulated with columns)
    writer.writerow([f"[LOGO: {COMPANY_NAME}]", "", "", "", f"Report: {title}"])
    writer.writerow(["", "", "", "", f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([COMPANY_TAGLINE, "", "", "", ""])
    writer.writerow([])  # Empty row
    
    # Column headers
    writer.writerow([col['label'] for col in columns])
    
    # Data rows
    for row in data:
        row_values = []
        for col in columns:
            value = row.get(col['key'], '')
            # Format currency columns
            if col.get('currency') and value:
                value = format_currency(value)
            row_values.append(value)
        writer.writerow(row_values)
    
    # Footer with totals - Label left, Value right
    if footer_data:
        writer.writerow([])  # Empty row
        writer.writerow(["=" * 40, "", "", "", "=" * 40])  # Separator
        for key, value in footer_data.items():
            # Create row with label on left and value on right
            empty_cols = [''] * (len(columns) - 2) if len(columns) > 2 else []
            writer.writerow([key] + empty_cols + [value])
    
    content = output.getvalue()
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
    )


def create_xlsx_response(data: list, columns: list, filename: str, title: str = "", footer_data: dict = None):
    """Create Excel file from data with header, logo, and footer"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
    title_font = Font(bold=True, size=14, color="3B5998")
    subtitle_font = Font(size=10, color="666666")
    currency_font = Font(name='Consolas', size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    footer_label_font = Font(bold=True, size=10)
    footer_value_font = Font(bold=True, size=10, color="3B5998")
    footer_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    current_row = 1
    last_col = len(columns)
    
    # Try to add logo on the LEFT
    logo_end_col = 3  # Logo spans columns A-C
    if LOGO_PATH.exists():
        try:
            img = Image(str(LOGO_PATH))
            img.width = 180
            img.height = 70
            ws.add_image(img, 'A1')
        except:
            pass
    
    # Company info on the RIGHT
    info_start_col = last_col - 1 if last_col > 3 else 4
    
    # Report title - right aligned
    cell = ws.cell(row=1, column=info_start_col, value=f"Report: {title}")
    cell.font = title_font
    cell.alignment = Alignment(horizontal='right')
    
    cell = ws.cell(row=2, column=info_start_col, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal='right')
    
    cell = ws.cell(row=3, column=info_start_col, value=COMPANY_TAGLINE)
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal='right')
    
    current_row = 6  # Leave space for logo
    
    # Column headers
    header_row = current_row
    for col_num, column in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_num, value=column['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    current_row += 1
    
    # Data rows
    for row_data in data:
        for col_num, column in enumerate(columns, 1):
            value = row_data.get(column['key'], '')
            cell = ws.cell(row=current_row, column=col_num)
            
            # Format currency columns
            if column.get('currency') and value:
                cell.value = format_currency(value)
                cell.font = currency_font
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = value
            
            cell.border = thin_border
        current_row += 1
    
    # Footer with totals - Label LEFT, Value RIGHT
    if footer_data:
        current_row += 1  # Empty row
        
        for key, value in footer_data.items():
            # Label on column 1, aligned left
            label_cell = ws.cell(row=current_row, column=1, value=key)
            label_cell.font = footer_label_font
            label_cell.fill = footer_fill
            label_cell.alignment = Alignment(horizontal='left')
            label_cell.border = thin_border
            
            # Merge middle columns for visual
            for mid_col in range(2, last_col):
                mid_cell = ws.cell(row=current_row, column=mid_col, value="")
                mid_cell.fill = footer_fill
                mid_cell.border = thin_border
            
            # Value on last column, aligned right
            value_cell = ws.cell(row=current_row, column=last_col, value=str(value))
            value_cell.font = footer_value_font
            value_cell.fill = footer_fill
            value_cell.alignment = Alignment(horizontal='right')
            value_cell.border = thin_border
            
            current_row += 1
    
    # Auto-adjust column widths
    for col_num, column in enumerate(columns, 1):
        max_length = len(column['label'])
        for row in ws.iter_rows(min_row=header_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )


def create_pdf_response(data: list, columns: list, title: str, filename: str, footer_data: dict = None):
    """Create PDF file from data with header, logo, and footer"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
    except ImportError:
        # Fallback to CSV if reportlab not available
        return create_csv_response(data, columns, filename, title, footer_data)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#3B5998'),
        alignment=2  # Right align
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.gray,
        alignment=2  # Right align
    )
    
    # Create header table with logo LEFT and info RIGHT
    header_data = []
    
    # Try to add logo
    logo_cell = ""
    if LOGO_PATH.exists():
        try:
            logo_cell = RLImage(str(LOGO_PATH), width=2.5*inch, height=1*inch)
        except:
            logo_cell = Paragraph(COMPANY_NAME, styles['Heading1'])
    else:
        logo_cell = Paragraph(COMPANY_NAME, styles['Heading1'])
    
    # Info on right
    info_text = f"""
    <para align="right">
    <font size="14" color="#3B5998"><b>Report: {title}</b></font><br/>
    <font size="9" color="#666666">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font><br/>
    <font size="9" color="#666666">{COMPANY_TAGLINE}</font>
    </para>
    """
    info_cell = Paragraph(info_text, styles['Normal'])
    
    header_table = Table([[logo_cell, info_cell]], colWidths=[3*inch, 7*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 20))
    
    # Prepare data table
    table_data = [[col['label'] for col in columns]]
    for row in data:
        row_values = []
        for col in columns:
            value = row.get(col['key'], '')
            if col.get('currency') and value:
                value = format_currency(value)
            row_values.append(str(value)[:40])  # Truncate long values
        table_data.append(row_values)
    
    table = Table(table_data, repeatRows=1)
    
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B5998')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
    ]))
    
    elements.append(table)
    
    # Footer with totals - Label LEFT, Value RIGHT
    if footer_data:
        elements.append(Spacer(1, 15))
        
        footer_table_data = []
        for key, value in footer_data.items():
            footer_table_data.append([key, str(value)])
        
        footer_table = Table(footer_table_data, colWidths=[4*inch, 2*inch])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F0F0')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),  # Labels bold
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),  # Values bold
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#3B5998')),  # Values in blue
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),   # Labels LEFT
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Values RIGHT
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        
        elements.append(footer_table)
    
    doc.build(elements)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
    )


@router.get("/movements")
async def export_movements(
    format: str = Query("csv", pattern="^(xlsx|pdf|csv)$"),
    movement_type: str = None,
    warehouse_id: str = None,
    date_from: str = None,
    date_to: str = None,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export stock movements to XLS, PDF or CSV"""
    query = select(models_ledger.StockMovement)\
        .where(models_ledger.StockMovement.tenant_id == current_user.tenant_id)\
        .options(
            selectinload(models_ledger.StockMovement.product),
            selectinload(models_ledger.StockMovement.location).selectinload(Location.warehouse),
            selectinload(models_ledger.StockMovement.user)
        )
    
    if movement_type:
        try:
            mt = models_ledger.MovementType(movement_type)
            query = query.where(models_ledger.StockMovement.movement_type == mt)
        except:
            pass
    
    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            query = query.where(models_ledger.StockMovement.timestamp >= dt_from)
        except:
            pass
    
    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            query = query.where(models_ledger.StockMovement.timestamp <= dt_to)
        except:
            pass
    
    result = await db.execute(query.order_by(models_ledger.StockMovement.timestamp.desc()).limit(500))
    movements = result.scalars().all()
    
    columns = [
        {'key': 'reference', 'label': 'Reference'},
        {'key': 'type', 'label': 'Type'},
        {'key': 'item', 'label': 'Item'},
        {'key': 'quantity', 'label': 'Quantity'},
        {'key': 'warehouse', 'label': 'Warehouse'},
        {'key': 'timestamp', 'label': 'Date/Time'},
        {'key': 'user', 'label': 'By'}
    ]
    
    data = []
    total_in = 0
    total_out = 0
    
    for m in movements:
        qty = m.quantity_change or 0
        if qty > 0:
            total_in += qty
        else:
            total_out += abs(qty)
        
        data.append({
            'reference': m.reference_id or f"MOV-{str(m.id)[:8]}",
            'type': m.movement_type.value if m.movement_type else "Unknown",
            'item': m.product.name if m.product else "Unknown",
            'quantity': qty,
            'warehouse': m.location.warehouse.name if m.location and m.location.warehouse else "Unknown",
            'timestamp': m.timestamp.strftime('%Y-%m-%d %H:%M') if m.timestamp else "",
            'user': m.user.username if m.user else "System"
        })
    
    footer_data = {
        'Total Records': len(data),
        'Total Inbound Qty': f"{total_in:,.0f}",
        'Total Outbound Qty': f"{total_out:,.0f}",
        'Net Movement': f"{total_in - total_out:,.0f}"
    }
    
    filename = f"movements_{datetime.now().strftime('%Y%m%d')}"
    title = "Stock Movements Report"
    
    if format == "xlsx":
        return create_xlsx_response(data, columns, filename, title, footer_data)
    elif format == "pdf":
        return create_pdf_response(data, columns, title, filename, footer_data)
    else:
        return create_csv_response(data, columns, filename, title, footer_data)


@router.get("/stock")
async def export_stock(
    format: str = Query("csv", pattern="^(xlsx|pdf|csv)$"),
    origin_type: str = None,
    warehouse_id: str = None,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export stock status to XLS, PDF or CSV"""
    query = select(InventoryBatch)\
        .where(
            InventoryBatch.tenant_id == current_user.tenant_id,
            InventoryBatch.quantity_on_hand > 0
        )\
        .options(
            selectinload(InventoryBatch.product),
            selectinload(InventoryBatch.location).selectinload(Location.warehouse)
        )
    
    if origin_type:
        from models.models_inventory import OriginType
        try:
            ot = OriginType(origin_type)
            query = query.where(InventoryBatch.origin_type == ot)
        except:
            pass
    
    result = await db.execute(query.limit(500))
    batches = result.scalars().all()
    
    columns = [
        {'key': 'product', 'label': 'Product'},
        {'key': 'code', 'label': 'Code'},
        {'key': 'batch', 'label': 'Batch'},
        {'key': 'quantity', 'label': 'Qty'},
        {'key': 'warehouse', 'label': 'Warehouse'},
        {'key': 'location', 'label': 'Location'},
        {'key': 'origin', 'label': 'Origin'},
        {'key': 'cost', 'label': 'Unit Cost', 'currency': True},
        {'key': 'total_value', 'label': 'Total Value', 'currency': True},
        {'key': 'expires', 'label': 'Expires'}
    ]
    
    data = []
    total_qty = 0
    total_value = 0
    
    for b in batches:
        qty = b.quantity_on_hand or 0
        cost = b.unit_cost or 0
        value = qty * cost
        total_qty += qty
        total_value += value
        
        data.append({
            'product': b.product.name if b.product else "Unknown",
            'code': b.product.code if b.product else "",
            'batch': b.batch_number,
            'quantity': qty,
            'warehouse': b.location.warehouse.name if b.location and b.location.warehouse else "Unknown",
            'location': b.location.name if b.location else "Unknown",
            'origin': b.origin_type.value if b.origin_type else "Unknown",
            'cost': cost,
            'total_value': value,
            'expires': b.expiration_date.strftime('%Y-%m-%d') if b.expiration_date else "N/A"
        })
    
    footer_data = {
        'Total Items': len(data),
        'Total Quantity': f"{total_qty:,.0f}",
        'Total Inventory Value': format_currency(total_value)
    }
    
    filename = f"stock_{datetime.now().strftime('%Y%m%d')}"
    title = "Stock Status Report"
    
    if format == "xlsx":
        return create_xlsx_response(data, columns, filename, title, footer_data)
    elif format == "pdf":
        return create_pdf_response(data, columns, title, filename, footer_data)
    else:
        return create_csv_response(data, columns, filename, title, footer_data)


@router.get("/products")
async def export_products(
    format: str = Query("csv", pattern="^(xlsx|pdf|csv)$"),
    product_type: str = None,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export products to XLS, PDF or CSV"""
    from models.models_manufacturing import Product
    
    query = select(Product).where(Product.tenant_id == current_user.tenant_id)
    
    if product_type:
        query = query.where(Product.type == product_type)
    
    result = await db.execute(query.limit(500))
    products = result.scalars().all()
    
    columns = [
        {'key': 'code', 'label': 'Code'},
        {'key': 'name', 'label': 'Name'},
        {'key': 'type', 'label': 'Type'},
        {'key': 'uom', 'label': 'UoM'},
        {'key': 'standard_cost', 'label': 'Standard Cost', 'currency': True},
        {'key': 'selling_price', 'label': 'Selling Price', 'currency': True},
        {'key': 'margin', 'label': 'Margin %'},
        {'key': 'manufactured', 'label': 'Manufactured'}
    ]
    
    data = []
    total_products = 0
    manufactured_count = 0
    purchased_count = 0
    
    for p in products:
        total_products += 1
        if p.is_manufactured:
            manufactured_count += 1
        else:
            purchased_count += 1
        
        cost = p.standard_cost or 0
        price = p.suggested_selling_price or 0
        margin = round((price - cost) / price * 100, 1) if price > 0 else 0
        
        data.append({
            'code': p.code,
            'name': p.name,
            'type': p.type.value if p.type else "Unknown",
            'uom': p.uom,
            'standard_cost': cost,
            'selling_price': price,
            'margin': f"{margin}%",
            'manufactured': 'Yes' if p.is_manufactured else 'No'
        })
    
    footer_data = {
        'Total Products': total_products,
        'Manufactured Products': manufactured_count,
        'Purchased Products': purchased_count
    }
    
    filename = f"products_{datetime.now().strftime('%Y%m%d')}"
    title = "Products Catalog Report"
    
    if format == "xlsx":
        return create_xlsx_response(data, columns, filename, title, footer_data)
    elif format == "pdf":
        return create_pdf_response(data, columns, title, filename, footer_data)
    else:
        return create_csv_response(data, columns, filename, title, footer_data)


@router.get("/production-orders")
async def export_production_orders(
    format: str = Query("csv", pattern="^(xlsx|pdf|csv)$"),
    status: str = None,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export production orders to XLS, PDF or CSV"""
    from models.models_manufacturing import ProductionOrder
    
    query = select(ProductionOrder).where(ProductionOrder.tenant_id == current_user.tenant_id)
    
    if status:
        query = query.where(ProductionOrder.status == status)
    
    result = await db.execute(query.order_by(ProductionOrder.scheduled_date.desc().nullslast()).limit(500))
    orders = result.scalars().all()
    
    columns = [
        {'key': 'order_no', 'label': 'Order No'},
        {'key': 'status', 'label': 'Status'},
        {'key': 'target_qty', 'label': 'Target Qty'},
        {'key': 'completed_qty', 'label': 'Completed'},
        {'key': 'progress', 'label': 'Progress'},
        {'key': 'material_cost', 'label': 'Material Cost', 'currency': True},
        {'key': 'labor_cost', 'label': 'Labor Cost', 'currency': True},
        {'key': 'total_hpp', 'label': 'Total HPP', 'currency': True},
        {'key': 'deadline', 'label': 'Deadline'}
    ]
    
    data = []
    total_target = 0
    total_completed = 0
    total_material = 0
    total_labor = 0
    total_hpp = 0
    
    for o in orders:
        target = o.target_qty or 0
        completed = o.completed_qty or 0
        material = o.material_cost or 0
        labor = o.labor_cost or 0
        hpp = o.total_hpp or 0
        
        total_target += target
        total_completed += completed
        total_material += material
        total_labor += labor
        total_hpp += hpp
        
        progress = round(completed / target * 100, 1) if target else 0
        
        data.append({
            'order_no': o.order_no,
            'status': o.status.value if hasattr(o.status, 'value') else str(o.status),
            'target_qty': target,
            'completed_qty': completed,
            'progress': f"{progress}%",
            'material_cost': material,
            'labor_cost': labor,
            'total_hpp': hpp,
            'deadline': o.deadline.strftime('%Y-%m-%d') if o.deadline else "N/A"
        })
    
    overall_progress = round(total_completed / total_target * 100, 1) if total_target else 0
    
    footer_data = {
        'Total Orders': len(data),
        'Total Target Qty': f"{total_target:,.0f}",
        'Total Completed': f"{total_completed:,.0f}",
        'Overall Progress': f"{overall_progress}%",
        'Total Material Cost': format_currency(total_material),
        'Total Labor Cost': format_currency(total_labor),
        'Total HPP': format_currency(total_hpp)
    }
    
    filename = f"production_orders_{datetime.now().strftime('%Y%m%d')}"
    title = "Production Orders Report"
    
    if format == "xlsx":
        return create_xlsx_response(data, columns, filename, title, footer_data)
    elif format == "pdf":
        return create_pdf_response(data, columns, title, filename, footer_data)
    else:
        return create_csv_response(data, columns, filename, title, footer_data)


# ============ Procurement Exports ============

@router.get("/vendors")
async def export_vendors(
    format: str = Query("csv", pattern="^(xlsx|pdf|csv)$"),
    rating: str = None,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export vendors to XLS, PDF or CSV"""
    from models.models_procurement import Vendor
    from datetime import datetime
    
    # Build query
    query = select(Vendor).where(Vendor.tenant_id == current_user.tenant_id)
    if rating:
        query = query.where(Vendor.rating == rating)
    query = query.order_by(Vendor.name)
    
    result = await db.execute(query)
    vendors = result.scalars().all()
    
    # Get currency settings
    currency_settings = await get_tenant_currency_settings(db, current_user.tenant_id)
    
    # Prepare data
    data = []
    for v in vendors:
        data.append({
            "code": v.code,
            "name": v.name,
            "phone": v.phone or "-",
            "email": v.email or "-",
            "rating": v.rating.value if v.rating else "-",
            "category": v.category.value if v.category else "-",
            "payment_term": v.payment_term.value if hasattr(v, 'payment_term') and v.payment_term else "Net 30",
            "credit_limit": format_currency(
                getattr(v, 'credit_limit', 0) or 0,
                currency_settings["symbol"],
                currency_settings["position"],
                currency_settings["thousand_sep"],
                currency_settings["decimal_sep"],
                int(currency_settings["decimal_places"])
            )
        })
    
    columns = [
        {"key": "code", "label": "Code"},
        {"key": "name", "label": "Vendor Name"},
        {"key": "phone", "label": "Phone"},
        {"key": "email", "label": "Email"},
        {"key": "rating", "label": "Rating"},
        {"key": "category", "label": "Category"},
        {"key": "payment_term", "label": "Payment Term"},
        {"key": "credit_limit", "label": "Credit Limit"}
    ]
    
    title = "Vendor List"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"vendors_{timestamp}.{format}"
    
    footer_data = {"Total Vendors": str(len(data))}
    
    if format == "xlsx":
        return create_xlsx_response(data, columns, filename, title, footer_data)
    elif format == "pdf":
        return create_pdf_response(data, columns, title, filename, footer_data)
    else:
        return create_csv_response(data, columns, filename, title, footer_data)


@router.get("/purchase-orders")
async def export_purchase_orders(
    format: str = Query("csv", pattern="^(xlsx|pdf|csv)$"),
    status: str = None,
    vendor_id: str = None,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export purchase orders to XLS, PDF or CSV"""
    from models.models_procurement import PurchaseOrder, Vendor
    from sqlalchemy.orm import selectinload
    from datetime import datetime
    
    # Build query
    query = select(PurchaseOrder).where(
        PurchaseOrder.tenant_id == current_user.tenant_id
    ).options(selectinload(PurchaseOrder.vendor))
    
    if status:
        query = query.where(PurchaseOrder.status == status)
    if vendor_id:
        query = query.where(PurchaseOrder.vendor_id == vendor_id)
    
    query = query.order_by(PurchaseOrder.created_at.desc())
    
    result = await db.execute(query)
    orders = result.scalars().all()
    
    # Get currency settings
    currency_settings = await get_tenant_currency_settings(db, current_user.tenant_id)
    
    # Calculate totals
    total_amount = sum(po.total_amount or 0 for po in orders)
    total_paid = sum(getattr(po, 'amount_paid', 0) or 0 for po in orders)
    
    # Prepare data
    data = []
    for po in orders:
        progress = getattr(po, 'progress', 0) or 0
        data.append({
            "po_number": po.po_number,
            "vendor": po.vendor.name if po.vendor else "-",
            "date": po.created_at.strftime("%Y-%m-%d") if po.created_at else "-",
            "status": po.status.value if po.status else "-",
            "payment_term": getattr(po, 'payment_term', None).value if hasattr(po, 'payment_term') and po.payment_term else "Net 30",
            "payment_status": getattr(po, 'payment_status', None).value if hasattr(po, 'payment_status') and po.payment_status else "Unpaid",
            "progress": f"{progress:.0f}%",
            "total_amount": format_currency(
                po.total_amount or 0,
                currency_settings["symbol"],
                currency_settings["position"],
                currency_settings["thousand_sep"],
                currency_settings["decimal_sep"],
                int(currency_settings["decimal_places"])
            ),
            "amount_paid": format_currency(
                getattr(po, 'amount_paid', 0) or 0,
                currency_settings["symbol"],
                currency_settings["position"],
                currency_settings["thousand_sep"],
                currency_settings["decimal_sep"],
                int(currency_settings["decimal_places"])
            )
        })
    
    columns = [
        {"key": "po_number", "label": "PO Number"},
        {"key": "vendor", "label": "Vendor"},
        {"key": "date", "label": "Date"},
        {"key": "status", "label": "Status"},
        {"key": "payment_term", "label": "Payment Term"},
        {"key": "payment_status", "label": "Payment Status"},
        {"key": "progress", "label": "Progress"},
        {"key": "total_amount", "label": "Total Amount"},
        {"key": "amount_paid", "label": "Amount Paid"}
    ]
    
    title = "Purchase Orders"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"purchase_orders_{timestamp}.{format}"
    
    footer_data = {
        "Total Orders": str(len(data)),
        "Total Amount": format_currency(
            total_amount,
            currency_settings["symbol"],
            currency_settings["position"],
            currency_settings["thousand_sep"],
            currency_settings["decimal_sep"],
            int(currency_settings["decimal_places"])
        ),
        "Total Paid": format_currency(
            total_paid,
            currency_settings["symbol"],
            currency_settings["position"],
            currency_settings["thousand_sep"],
            currency_settings["decimal_sep"],
            int(currency_settings["decimal_places"])
        )
    }
    
    if format == "xlsx":
        return create_xlsx_response(data, columns, filename, title, footer_data)
    elif format == "pdf":
        return create_pdf_response(data, columns, title, filename, footer_data)
    else:
        return create_csv_response(data, columns, filename, title, footer_data)


# ============ Stock Opname Exports ============

@router.get("/opnames")
async def export_opnames(
    format: str = Query("csv", pattern="^(xlsx|pdf|csv)$"),
    status: str = None,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export stock opnames to XLS, PDF or CSV"""
    from sqlalchemy.orm import selectinload
    
    query = select(models.StockOpname).where(
        models.StockOpname.tenant_id == current_user.tenant_id
    ).options(
        selectinload(models.StockOpname.warehouse),
        selectinload(models.StockOpname.details)
    ).order_by(models.StockOpname.date.desc())
    
    if status:
        query = query.where(models.StockOpname.status == status.upper())
    
    result = await db.execute(query)
    opnames = result.scalars().all()
    
    # Format data
    data = []
    for opname in opnames:
        # Count differences
        total_items = len(opname.details) if opname.details else 0
        items_with_diff = 0
        total_diff = 0
        if opname.details:
            for d in opname.details:
                if d.counted_qty is not None:
                    diff = d.counted_qty - d.system_qty
                    total_diff += diff
                    if diff != 0:
                        items_with_diff += 1
        
        data.append({
            'date': opname.date.strftime("%Y-%m-%d") if opname.date else "",
            'warehouse': opname.warehouse.name if opname.warehouse else "-",
            'status': opname.status.value if hasattr(opname.status, 'value') else str(opname.status),
            'total_items': total_items,
            'items_with_diff': items_with_diff,
            'net_adjustment': total_diff,
            'notes': opname.notes or ""
        })
    
    columns = [
        {'key': 'date', 'label': 'Date'},
        {'key': 'warehouse', 'label': 'Warehouse'},
        {'key': 'status', 'label': 'Status'},
        {'key': 'total_items', 'label': 'Total Items'},
        {'key': 'items_with_diff', 'label': 'Items w/ Diff'},
        {'key': 'net_adjustment', 'label': 'Net Adjustment'},
        {'key': 'notes', 'label': 'Notes'}
    ]
    
    title = "Stock Opname Report"
    filename = f"stock_opname_{datetime.now().strftime('%Y%m%d')}"
    
    # Footer data
    footer_data = {
        'Total Records': len(data),
        'Total Adjustments': sum(d.get('net_adjustment', 0) for d in data)
    }
    
    if format == "xlsx":
        return create_xlsx_response(data, columns, filename, title, footer_data)
    elif format == "pdf":
        return create_pdf_response(data, columns, title, filename, footer_data)
    else:
        return create_csv_response(data, columns, filename, title, footer_data)
